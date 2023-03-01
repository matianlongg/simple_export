# -*- coding: utf-8 -*-
# !/usr/bin/env python
# @Time    : 2022/11/9 15:03 
# @Author  : mtl
# @Desc    : excel 根据 模板导出工具
# @File    : index.py
# @Software: PyCharm
import collections
import copy
import io
import traceback
from pathlib import Path

from PIL import Image
from openpyxl.formatting.formatting import ConditionalFormatting
from openpyxl.workbook import Workbook
from openpyxl.cell import Cell, MergedCell, cell
import re
from openpyxl.worksheet.worksheet import Worksheet
from simple_export.utils.tool import to_flat, pos_char_to_num, num_to_pos_char, char_to_num, points_to_pixels
from openpyxl.drawing.image import Image as oImage
from simple_export.utils.util import dynamic_method
"""
基于openpyxl 根据字典数据导出
"""

class work_sheet_tool():
    def __init__(self, convert_pic=False, dynamic_method=dynamic_method()):
        self.convert_pic = convert_pic
        self.active_sheet: Worksheet = None
        self.active_source_sheet: Worksheet = None
        self.task_queue = collections.deque()
        self.dynamic_method = dynamic_method
        self.col_queue = collections.deque()

    def copy_cell(self, source_cell: Cell, target_cell: Cell) -> None:
        """
        复制cell
        :param source_cell: 源cell
        :param target_cell: 目标cell
        :return:
        """
        target_cell._style = copy.copy(source_cell._style)
        target_cell.font = copy.copy(source_cell.font)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.number_format = copy.copy(source_cell.number_format)
        target_cell.protection = copy.copy(source_cell.protection)
        target_cell.alignment = copy.copy(source_cell.alignment)
        self.write_val(target_cell, source_cell.value)

    def write_attr(self, source: Worksheet, target: Worksheet, pos_mapping: dict):
        """
        写入源sheet的一些属性
        :param source_cell: 源sheet
        :param target_cell: 目标sheet
        :return:
        """
        target.views.sheetView = source.views.sheetView
        target._rels = source._rels
        target._drawing = source._drawing
        con: ConditionalFormatting
        for con in source.conditional_formatting:
            for cell in con.cells:
                left_top: list
                right_bottom: list
                if ":" in cell.coord:
                    left_top, right_bottom = pos_char_to_num(cell.coord)
                    right_bottom = (right_bottom[0], pos_mapping[right_bottom[1] - 1][-1])
                else:
                    x, y = char_to_num(cell.coord)
                    left_top, right_bottom = [x, y], [x, pos_mapping[y - 1][-1]]

                for rule in con.rules:
                    target.conditional_formatting.add(f"{num_to_pos_char(left_top)}:{num_to_pos_char(right_bottom)}", rule)
        target.data_validations = source.data_validations

    def write_val(self, target_cell: Cell, val):
        """
        写入值
        :param target_cell: 写入单元格
        :param val: 写入单元格的值
        :return:
        """
        if self.convert_pic and isinstance(val, str):
            file_name = Path(val)
            if val.startswith("http"):
                import requests
                headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.75 Safari/537.36'}
                req = requests.get(headers=headers, url=val)
                ximg = Image.open(io.BytesIO(req.content))
                self.write_img(target_cell, ximg)
                return
            elif file_name.suffix.lower() in ['.jpg', '.png', '.gif']:
                if file_name.is_file():
                    ximg = Image.open(val)
                    self.write_img(target_cell, ximg)
                    return
        target_cell.value = val

    def write_img(self, target_cell: Cell, ximg: Image):
        """
        写入图片
        :param target_cell: 写入的cell
        :param ximg: 待写入的图片
        :return:
        """
        w, h = ximg.size
        w_h_ratio = w / h
        width: float = 30
        if target_cell.column_letter in self.active_sheet.column_dimensions:
            width = self.active_sheet.column_dimensions[target_cell.column_letter].width
        height: float = self.active_sheet.row_dimensions[target_cell.row].height
        if width is None:
            width = 30
        if height is None:
            height = 30
        height = int(points_to_pixels(height) * 0.9)
        img_width = int(width * 8)
        width = int(w_h_ratio * height * 0.9)
        if width > img_width:
            width = img_width
        ximg = ximg.resize((width, height))
        aimg = oImage(ximg)
        aimg.format = "jpg"
        self.active_sheet.add_image(aimg, target_cell.coordinate)

    def get_task(self, coordinate: str, val: str) -> str:
        if "|" in val:
            vals: [str] = val.split("|")
            self.task_queue.append((vals[-1].split(";"), coordinate))
            return vals[0]
        return val

    def exec_task(self, target, pos_mapping: collections.defaultdict, pos: collections.defaultdict):
        while self.task_queue:
            func, coordinate = self.task_queue.popleft()
            self.dynamic_method(func, coordinate, target, pos_mapping, pos)

    def replace_pos(self, start: int, col: int, pos: list):
        n = start
        for i in range(start + 1, len(pos)):
            if pos[i][col] != 0:
                pos[i][col], pos[n][col] = pos[n][col], pos[i][col]
                n = i

    def is_exce(self, col: Cell):
        val = col.value
        if isinstance(val, str):
            match_value: list = re.findall('\${(.+)}', col.value)
            return match_value
        return []
    def write_sheet(self, source: Worksheet, obj_value: dict, target: Worksheet):
        pos = []
        self.col_queue.clear()
        pos_mapping: collections.defaultdict = collections.defaultdict(list)
        for i, row in enumerate(source.iter_rows()):
            pos.append(len(row) * [0])
            for j, col in enumerate(row):
                for value in self.is_exce(col):
                    match_value_str: str = self.get_task(col.coordinate, value)
                    index: int = match_value_str.find("*")
                    # 找到 ${}
                    n = 0
                    if index > 0:
                        key: str = f"{match_value_str[:index]}%s{match_value_str[index + 1:]}"
                        okey = key % n
                        if okey not in obj_value.keys():
                            col.value = ""
                            break
                        col.value = obj_value[okey]
                        self.col_queue.append((i + 1, j, n + 1, key))
                    else:
                        col.value = obj_value.get(match_value_str, "")
                pos[-1][j] = col

            if self.col_queue:
                pos_mapping[i] = self.write_list(pos, obj_value)
            else:
                pos_mapping[i] = [len(pos)]
        target.insert_rows(len(pos))
        for index in source.column_dimensions:
            target.column_dimensions[index].width = source.column_dimensions[index].width
        for index in source.row_dimensions:
            height = source.row_dimensions[index].height
            if height is None:
                height = 30
            for pm in pos_mapping[index - 1]:
                target.row_dimensions[pm].height = height
        for i, r in enumerate(pos):
            for j, c in enumerate(r):
                source_cell: [Cell, MergedCell] = pos[i][j]
                target_cell: Cell = target.cell(i + 1, j + 1)
                self.copy_cell(source_cell, target_cell)
        for i, cell in enumerate(source.merged_cells):
            try:
                cell_min_num = pos_mapping[cell.min_row - 1][0]
                cell_max_num = pos_mapping[cell.max_row - 1][-1]
                target.merge_cells(start_row=cell_min_num, start_column=cell.min_col,
                                   end_column=cell.max_col, end_row=cell_max_num)
            except:
                traceback.print_exc()
        self.exec_task(target, pos_mapping, pos)
        self.write_table(source, target, pos_mapping)
        self.write_attr(source, target, pos_mapping)

    def write_list(self, pos: list, obj_value: dict) -> list:
        tmp_row: list[Cell] = [copy.copy(col) for col in pos[-1]]
        col_list = []
        insert_pos = [len(pos)]
        while self.col_queue:
            i, j, n, key = self.col_queue.popleft()
            okey = key % n
            if okey not in obj_value.keys():
                continue
            tmp_row[j].value = obj_value[okey]
            col_list.append((i + 1, j, n + 1, key))
        if len(col_list) > 0:
            pos.append(tmp_row)
            self.col_queue.extend(col_list)
            insert_pos.extend(self.write_list(pos, obj_value))
        return insert_pos

    def write_shee2t(self, source: Worksheet, obj_value: dict, target: Worksheet):
        """
        查找${}里的值 跟obj_value进行比对、替换
        :param source: 源sheet
        :param obj_value: 传入的值 {}
        :param target: 目标sheet
        :return:
        """
        pos: list = [list(row) for row in source.iter_rows()]
        pos_mapping: collections.defaultdict = collections.defaultdict(dict)
        rlen: int = len(pos)
        clen: int = len(pos[0])
        i: int = 0
        w: int = 0
        while i < rlen:
            j: int = 0
            maxn: int = 0
            while j < clen:
                try:
                    if pos[i][j] == 0:
                        self.replace_pos(i, j, pos)
                    if j not in pos_mapping[i]:
                        pos_mapping[i].setdefault(j, []).append(
                            pos_mapping[i - 1][j][-1] + 1 if i - 1 > 0 else i
                        )
                    if isinstance(pos[i][j], Cell) and isinstance(pos[i][j].value, str):
                        col: Cell = pos[i][j]
                        match_value: list = re.findall('\${(.+)}', col.value)
                        if len(match_value) > 0:
                            match_value_str: str = self.get_task(col.coordinate, match_value[0])
                            index: int = match_value_str.find("*")
                            n: int = 0
                            if index > 0:
                                while True:
                                    key: str = f"{match_value_str[:index]}{n}{match_value_str[index + 1:]}"
                                    if key in obj_value.keys():
                                        if n + i not in pos_mapping[i][j]:
                                            pos_mapping[i][j].append(n + i)
                                        if n == 0:
                                            col.value = obj_value[key]
                                        elif n + i >= len(pos) or pos[n + i][j] != 0:
                                            ls_col = [0] * clen
                                            ccol = copy.copy(col)
                                            ccol.value = obj_value[key]
                                            ls_col[j] = ccol
                                            pos.insert(n + i, ls_col)
                                        else:
                                            ccol = copy.copy(col)
                                            ccol.value = obj_value[key]
                                            pos[n + i][j] = ccol

                                        n += 1
                                    else:
                                        col.value = ""
                                        break
                                maxn = max(n - 1, maxn)
                            else:
                                col.value = obj_value.get(match_value_str, "")
                except:
                    traceback.print_exc()
                j += 1
            if maxn > 0:
                rlen += maxn
                w += maxn
            i += 1
        target.insert_rows(len(pos))
        for index in source.column_dimensions:
            target.column_dimensions[index].width = source.column_dimensions[index].width
        for index in source.row_dimensions:
            for pm in max(list(pos_mapping[index - 1].values())):
                height = source.row_dimensions[index].height
                if height is None:
                    height = 30
                target.row_dimensions[pm + 1].height = height
        for i, r in enumerate(pos):
            for j, c in enumerate(r):
                if isinstance(pos[i][j], Cell):
                    source_cell: [Cell, MergedCell] = pos[i][j]
                    target_cell: Cell = target.cell(i + 1, j + 1)
                    self.copy_cell(source_cell, target_cell)
        for i, cell in enumerate(source.merged_cells):
            try:
                cell_min_num = pos_mapping[cell.min_row - 1][cell.min_col - 1][0] + 1
                cell_max_num = pos_mapping[cell.max_row - 1][cell.max_col - 1][0] + 1
                target.merge_cells(start_row=cell_min_num, start_column=cell.min_col,
                                end_column=cell.max_col, end_row=cell_max_num)
            except:
                traceback.print_exc()
        self.exec_task(target, pos_mapping, pos)
        self.write_table(source, target, pos_mapping)
        self.write_attr(source, target, pos_mapping)

    def write_table(self, source: Worksheet, target: Worksheet, pos_mapping: dict):
        """
        复制table
        :param source:
        :param target:
        :param pos_mapping: 源excel 和 经过处理的excel的行对应关系
        :return:
        """
        tab: tuple
        for tab in source.tables.items():
            left_top, right_bottom = pos_char_to_num(tab[1])
            left_top = (left_top[0], pos_mapping[left_top[1] - 1][0])
            right_bottom = (right_bottom[0], pos_mapping[right_bottom[1] - 1][-1])
            ctab = copy.deepcopy(source.tables[tab[0]])
            ctab.ref = f"{num_to_pos_char(left_top)}:{num_to_pos_char(right_bottom)}"
            ctab.autoFilter.ref = ctab.ref
            target.add_table(ctab)

def write_excel_for_template(value: dict, wb_tmp: Workbook, *args, **kwargs) -> None:
    """
    扁平value，查找和value一级key对应的sheet，处理
    :param value: 展示值
    :param wb_tmp: workbook
    :return:
    """
    sheet_name: str
    wst: work_sheet_tool = work_sheet_tool(*args, **kwargs)
    for sheet_name, obj in value.items():
        obj_value: dict = to_flat(obj)
        if sheet_name in wb_tmp.sheetnames:
            source: Worksheet = wb_tmp[sheet_name]
            target: Worksheet = wb_tmp.create_sheet("new_" + sheet_name, wb_tmp.index(source))
            wst.active_sheet = target
            wst.active_source_sheet = source
            wb_tmp.remove(source)
            wst.write_sheet(source, obj_value, target)
            if obj_value.get("sheet_name", ""):
                target.title = obj_value["sheet_name"]
            else:
                target.title = sheet_name