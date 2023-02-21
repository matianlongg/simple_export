#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2022/12/10 0:33
# @Author  : mtl
# @File    : example.py
# @Description : *****
import collections
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel import write_excel_for_template
from simple_export.utils.util import dynamic_method
class dynamic_method2(dynamic_method):
    def abc(self, coordinate: str, sheet: Worksheet, pos_mapping: collections.defaultdict, pos: collections.defaultdict ):
        print("abc")

    def get_sum(self, coordinate: str, sheet: Worksheet, pos_mapping: collections.defaultdict, pos: collections.defaultdict, a):
        print("abc")

def test1():
    wb_tmp = load_workbook(Path(__file__).parent / 'template/excel1.xlsx')
    value = {
        "家庭财产库存清单": {
            "thing": [
                {
                    "id": "=ROW($A1)",
                    "room": "客厅",
                    "goods": "物品1",
                    "structure": "制造商1",
                    "serial_number": "33XCBH3",
                    "time": "=TODAY()-120",
                    "source": "联机",
                    "price": 2000,
                    "eval_price": 2000,
                    "remark": "",
                    "has_photo": "是",
                    "photo_url": "https://t13.baidu.com/it/u=3512963775,847648221&fm=224&app=112&f=JPEG?w=500&h=500"
                },
                {
                    "id": "=ROW($A1)",
                    "room": "客厅2",
                    "goods": "物品2",
                    "structure": "制造商2",
                    "serial_number": "33XCBH4",
                    "time": "=TODAY()-90",
                    "source": "联机",
                    "price": 1000,
                    "eval_price": 1000,
                    "remark": "",
                    "has_photo": "是",
                    "photo_url": r"C:\Users\e9\Pictures\微信图片_20220616151258.png"
                }
            ],
            "thing2": [
                {
                    "id": "=ROW($A1)",
                    "room": "客厅",
                    "goods": "物品1",
                    "structure": "制造商1",
                    "serial_number": "33XCBH3",
                    "time": "=TODAY()-120",
                    "source": "联机",
                    "price": 2000,
                    "eval_price": 2000,
                    "remark": "",
                    "has_photo": "是",
                    "photo_url": "https://t13.baidu.com/it/u=3512963775,847648221&fm=224&app=112&f=JPEG?w=500&h=500"
                },
                {
                    "id": "=ROW($A1)",
                    "room": "客厅2",
                    "goods": "物品2",
                    "structure": "制造商2",
                    "serial_number": "33XCBH4",
                    "time": "=TODAY()-90",
                    "source": "联机",
                    "price": 1000,
                    "eval_price": 1000,
                    "remark": "",
                    "has_photo": "是",
                    "photo_url": r"C:\Users\e9\Pictures\微信图片_20220616151258.png"
                }
            ],
            "test": "1"
        },
        "sheet1": {
            "data": {
                "name": "1",
                "data_list": [
                    {"a": 3},
                    {"a": 4}
                ],
                "b": "cec测试"
            },
            "data_list": [
                {
                    "a": 1,
                    "b": 2
                },
                {
                    "a": 2,
                    "b": 1
                }
            ],
            "name": 123
        },
        "sheet2": {
            "data": {
                "name": "1",
                "data_list": [
                    {"a": 3}
                ],
                "b": "cec测试"
            }
        }
    }
    write_excel_for_template(value=value, wb_tmp=wb_tmp, convert_pic=True)
    wb_tmp.save("./val1.xlsx")
    wb_tmp.close()
    wb_tmp = load_workbook(f'./val1.xlsx', data_only=True, read_only=True)


def test2():
    wb_tmp = load_workbook(Path(__file__).parent / 'template/excel2.xlsx')
    value = {
        "sheet1": {
            "data": {
                "name": "1",
                "data_list": [
                    {"a": 3},
                    {"a": 4}
                ],
                "b": "cec测试"
            },
            "data_list": [
                {
                    "a": 1,
                    "b": 2
                },
                {
                    "a": 2,
                    "b": 1
                }
            ],
            "name": 123
        },
        "Sheet2": {
            "data": {
                "b": "1"
            },
            "b": "cec测试"
        }
    }
    write_excel_for_template(value=value, wb_tmp=wb_tmp, convert_pic=True)
    wb_tmp.save("./val2.xlsx")
    # wb_tmp.close()
    # wb_tmp = load_workbook(f'./vaL2.xlsx', data_only=True, read_only=True)

def test3():
    value = {'sheet1': {'row1': [
        {'outlet_id': 5, 'proce_id': 1, 'proce_name': '原料工序', 'production_id': 4, 'production_name': '一期烧结机',
         'outlet_name': '1#烧结机尾废气排放口', 'outlet_type': '7', 'factor': 'dust', 'criterion': None, 'tax_number': None,
         'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口', 'factor_name': '颗粒物(粉尘)', 'id': 3279,
         'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0, 'hour_emissions': '-', 'gas_emissions': '4864401.109',
         'pollutant_concentration': '0.0', 'pollutant_emissions': '158.76', 'pollutant_equivalent': '79.38',
         'tax': '158.76', 'discount': '50.0%', 'tax_discount': '79.38', 'tax_actual': '79.38', 'yield': '0.0',
         'product_pollute': '0.0', 'discharge_pollute': '0.0', 'Fe_consume': '-', 'Fe_S_content': '-',
         'solid_consume': '-', 'solid_S_content': '-', 'pollutant_yield': '0.0', 'pollutant_emissions_coeff': '0.0',
         'pollutant_yield_equivalent': '0.0', 'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0',
         'emissions_tax': '0.0', 'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None,
         'standard': None, 'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 7, 'proce_id': 1, 'proce_name': '原料工序', 'production_id': 5, 'production_name': '二期烧结机',
         'outlet_name': '2#烧结机尾废气排放口', 'outlet_type': '7', 'factor': 'dust', 'criterion': None, 'tax_number': None,
         'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口', 'factor_name': '颗粒物(粉尘)', 'id': 3280,
         'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0, 'hour_emissions': '-', 'gas_emissions': '2933315.636',
         'pollutant_concentration': None, 'pollutant_emissions': '168.23', 'pollutant_equivalent': '84.11499999999998',
         'tax': '168.23', 'discount': '0%', 'tax_discount': '0.0', 'tax_actual': '168.23', 'yield': '0.0',
         'product_pollute': '0.0', 'discharge_pollute': '0.0', 'Fe_consume': '-', 'Fe_S_content': '-',
         'solid_consume': '-', 'solid_S_content': '-', 'pollutant_yield': '0.0', 'pollutant_emissions_coeff': '0.0',
         'pollutant_yield_equivalent': '0.0', 'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0',
         'emissions_tax': '0.0', 'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None,
         'standard': None, 'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 8, 'proce_id': 2, 'proce_name': '炼铁工序', 'production_id': 6, 'production_name': '1#高炉',
         'outlet_name': '新1#高炉出铁场废气排放口', 'outlet_type': '7', 'factor': 'smoke', 'criterion': None, 'tax_number': None,
         'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口', 'factor_name': '颗粒物(烟尘)', 'id': 3281,
         'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0, 'hour_emissions': '-',
         'gas_emissions': '8916220.665000001', 'pollutant_concentration': None, 'pollutant_emissions': '177.25',
         'pollutant_equivalent': '88.625', 'tax': '88.62', 'discount': '0%', 'tax_discount': '0.0',
         'tax_actual': '88.62', 'yield': '0.0', 'product_pollute': '0.0', 'discharge_pollute': '0.0', 'Fe_consume': '-',
         'Fe_S_content': '-', 'solid_consume': '-', 'solid_S_content': '-', 'pollutant_yield': '0.0',
         'pollutant_emissions_coeff': '0.0', 'pollutant_yield_equivalent': '0.0',
         'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0', 'emissions_tax': '0.0',
         'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None, 'standard': None,
         'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 9, 'proce_id': 2, 'proce_name': '炼铁工序', 'production_id': 6, 'production_name': '1#高炉',
         'outlet_name': '1#高炉矿槽废气排放口（南）', 'outlet_type': '7', 'factor': 'dust', 'criterion': None, 'tax_number': None,
         'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口', 'factor_name': '颗粒物(粉尘)', 'id': 3282,
         'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0, 'hour_emissions': '-',
         'gas_emissions': '2501703.6950000003', 'pollutant_concentration': None, 'pollutant_emissions': '48.32',
         'pollutant_equivalent': '24.160000000000004', 'tax': '48.32', 'discount': '0%', 'tax_discount': '0.0',
         'tax_actual': '48.32', 'yield': '1.0', 'product_pollute': '0.0', 'discharge_pollute': '0.0', 'Fe_consume': '-',
         'Fe_S_content': '-', 'solid_consume': '-', 'solid_S_content': '-', 'pollutant_yield': '0.0',
         'pollutant_emissions_coeff': '0.0', 'pollutant_yield_equivalent': '0.0',
         'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0', 'emissions_tax': '0.0',
         'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None, 'standard': None,
         'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 10, 'proce_id': 2, 'proce_name': '炼铁工序', 'production_id': 6, 'production_name': '1#高炉',
         'outlet_name': '1#高炉矿槽废气排放口（北）', 'outlet_type': '7', 'factor': 'dust', 'criterion': None, 'tax_number': None,
         'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口', 'factor_name': '颗粒物(粉尘)', 'id': 3283,
         'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0, 'hour_emissions': '-', 'gas_emissions': '6780176.809',
         'pollutant_concentration': None, 'pollutant_emissions': '97.15', 'pollutant_equivalent': '48.575',
         'tax': '97.15', 'discount': '0%', 'tax_discount': '0.0', 'tax_actual': '97.15', 'yield': '1.0',
         'product_pollute': '0.0', 'discharge_pollute': '0.0', 'Fe_consume': '-', 'Fe_S_content': '-',
         'solid_consume': '-', 'solid_S_content': '-', 'pollutant_yield': '0.0', 'pollutant_emissions_coeff': '0.0',
         'pollutant_yield_equivalent': '0.0', 'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0',
         'emissions_tax': '0.0', 'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None,
         'standard': None, 'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 11, 'proce_id': 2, 'proce_name': '炼铁工序', 'production_id': 7, 'production_name': '2#高炉',
         'outlet_name': '2#高炉出铁场废气排放口', 'outlet_type': '7', 'factor': 'smoke', 'criterion': None, 'tax_number': None,
         'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口', 'factor_name': '颗粒物(烟尘)', 'id': 3284,
         'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0, 'hour_emissions': '-', 'gas_emissions': '5056214.688',
         'pollutant_concentration': None, 'pollutant_emissions': '51.25', 'pollutant_equivalent': '25.625',
         'tax': '25.62', 'discount': '0%', 'tax_discount': '0.0', 'tax_actual': '25.62', 'yield': '2.0',
         'product_pollute': '0.0', 'discharge_pollute': '0.0', 'Fe_consume': '-', 'Fe_S_content': '-',
         'solid_consume': '-', 'solid_S_content': '-', 'pollutant_yield': '0.0', 'pollutant_emissions_coeff': '0.0',
         'pollutant_yield_equivalent': '0.0', 'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0',
         'emissions_tax': '0.0', 'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None,
         'standard': None, 'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 12, 'proce_id': 2, 'proce_name': '炼铁工序', 'production_id': 7, 'production_name': '2#高炉',
         'outlet_name': '2#高炉矿槽废气排放口', 'outlet_type': '7', 'factor': 'dust', 'criterion': None, 'tax_number': None,
         'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口', 'factor_name': '颗粒物(粉尘)', 'id': 3285,
         'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0, 'hour_emissions': '-',
         'gas_emissions': '3558232.7970000003', 'pollutant_concentration': None, 'pollutant_emissions': '103.82',
         'pollutant_equivalent': '51.91', 'tax': '103.82', 'discount': '0%', 'tax_discount': '0.0',
         'tax_actual': '103.82', 'yield': '1.0', 'product_pollute': '0.0', 'discharge_pollute': '0.0',
         'Fe_consume': '-', 'Fe_S_content': '-', 'solid_consume': '-', 'solid_S_content': '-', 'pollutant_yield': '0.0',
         'pollutant_emissions_coeff': '0.0', 'pollutant_yield_equivalent': '0.0',
         'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0', 'emissions_tax': '0.0',
         'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None, 'standard': None,
         'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 13, 'proce_id': 3, 'proce_name': '炼钢工序', 'production_id': 8, 'production_name': '1#转炉',
         'outlet_name': '转炉二次烟气排放口', 'outlet_type': '7', 'factor': 'dust', 'criterion': None, 'tax_number': None,
         'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口', 'factor_name': '颗粒物(粉尘)', 'id': 3286,
         'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0, 'hour_emissions': '-', 'gas_emissions': '7921141.089',
         'pollutant_concentration': None, 'pollutant_emissions': '94.77', 'pollutant_equivalent': '47.38500000000001',
         'tax': '94.77', 'discount': '0%', 'tax_discount': '0.0', 'tax_actual': '94.77', 'yield': '0.0',
         'product_pollute': '0.0', 'discharge_pollute': '0.0', 'Fe_consume': '-', 'Fe_S_content': '-',
         'solid_consume': '-', 'solid_S_content': '-', 'pollutant_yield': '0.0', 'pollutant_emissions_coeff': '0.0',
         'pollutant_yield_equivalent': '0.0', 'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0',
         'emissions_tax': '0.0', 'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None,
         'standard': None, 'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 13, 'proce_id': 3, 'proce_name': '炼钢工序', 'production_id': 9, 'production_name': '2#转炉',
         'outlet_name': '转炉二次烟气排放口', 'outlet_type': '7', 'factor': 'dust', 'criterion': None, 'tax_number': None,
         'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口', 'factor_name': '颗粒物(粉尘)', 'id': 3286,
         'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0, 'hour_emissions': '-', 'gas_emissions': '7921141.089',
         'pollutant_concentration': None, 'pollutant_emissions': '94.77', 'pollutant_equivalent': '47.38500000000001',
         'tax': '94.77', 'discount': '0%', 'tax_discount': '0.0', 'tax_actual': '94.77', 'yield': '0.0',
         'product_pollute': '0.0', 'discharge_pollute': '0.0', 'Fe_consume': '-', 'Fe_S_content': '-',
         'solid_consume': '-', 'solid_S_content': '-', 'pollutant_yield': '0.0', 'pollutant_emissions_coeff': '0.0',
         'pollutant_yield_equivalent': '0.0', 'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0',
         'emissions_tax': '0.0', 'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None,
         'standard': None, 'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 15, 'proce_id': 1, 'proce_name': '原料工序', 'production_id': 4, 'production_name': '一期烧结机',
         'outlet_name': '烧结烟气脱硫脱硝废气排放口', 'outlet_type': '7', 'factor': 'smoke', 'criterion': None, 'tax_number': None,
         'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口', 'factor_name': '颗粒物(烟尘)', 'id': 3289,
         'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0, 'hour_emissions': '-',
         'gas_emissions': '12120451.946999999', 'pollutant_concentration': None, 'pollutant_emissions': '2253.2',
         'pollutant_equivalent': '751.0666666666667', 'tax': '2253.2', 'discount': '0%', 'tax_discount': '0.0',
         'tax_actual': '2253.2', 'yield': '0.0', 'product_pollute': '0.0', 'discharge_pollute': '0.0',
         'Fe_consume': '-', 'Fe_S_content': '-', 'solid_consume': '-', 'solid_S_content': '-', 'pollutant_yield': '0.0',
         'pollutant_emissions_coeff': '0.0', 'pollutant_yield_equivalent': '0.0',
         'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0', 'emissions_tax': '0.0',
         'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None, 'standard': None,
         'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 15, 'proce_id': 1, 'proce_name': '原料工序', 'production_id': 5, 'production_name': '二期烧结机',
         'outlet_name': '烧结烟气脱硫脱硝废气排放口', 'outlet_type': '7', 'factor': 'smoke', 'criterion': None, 'tax_number': None,
         'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口', 'factor_name': '颗粒物(烟尘)', 'id': 3289,
         'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0, 'hour_emissions': '-',
         'gas_emissions': '12120451.946999999', 'pollutant_concentration': None, 'pollutant_emissions': '2253.2',
         'pollutant_equivalent': '751.0666666666667', 'tax': '2253.2', 'discount': '0%', 'tax_discount': '0.0',
         'tax_actual': '2253.2', 'yield': '0.0', 'product_pollute': '0.0', 'discharge_pollute': '0.0',
         'Fe_consume': '-', 'Fe_S_content': '-', 'solid_consume': '-', 'solid_S_content': '-', 'pollutant_yield': '0.0',
         'pollutant_emissions_coeff': '0.0', 'pollutant_yield_equivalent': '0.0',
         'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0', 'emissions_tax': '0.0',
         'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None, 'standard': None,
         'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 15, 'proce_id': 1, 'proce_name': '原料工序', 'production_id': 4, 'production_name': '一期烧结机',
         'outlet_name': '烧结烟气脱硫脱硝废气排放口', 'outlet_type': '7', 'factor': 'nox', 'criterion': None, 'tax_number': None,
         'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口', 'factor_name': '氮氧化物', 'id': 3289,
         'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0, 'hour_emissions': '-',
         'gas_emissions': '12120451.946999999', 'pollutant_concentration': None, 'pollutant_emissions': '2253.2',
         'pollutant_equivalent': '751.0666666666667', 'tax': '2253.2', 'discount': '0%', 'tax_discount': '0.0',
         'tax_actual': '2253.2', 'yield': '0.0', 'product_pollute': '0.0', 'discharge_pollute': '0.0',
         'Fe_consume': '-', 'Fe_S_content': '-', 'solid_consume': '-', 'solid_S_content': '-', 'pollutant_yield': '0.0',
         'pollutant_emissions_coeff': '0.0', 'pollutant_yield_equivalent': '0.0',
         'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0', 'emissions_tax': '0.0',
         'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None, 'standard': None,
         'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 15, 'proce_id': 1, 'proce_name': '原料工序', 'production_id': 5, 'production_name': '二期烧结机',
         'outlet_name': '烧结烟气脱硫脱硝废气排放口', 'outlet_type': '7', 'factor': 'nox', 'criterion': None, 'tax_number': None,
         'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口', 'factor_name': '氮氧化物', 'id': 3289,
         'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0, 'hour_emissions': '-',
         'gas_emissions': '12120451.946999999', 'pollutant_concentration': None, 'pollutant_emissions': '2253.2',
         'pollutant_equivalent': '751.0666666666667', 'tax': '2253.2', 'discount': '0%', 'tax_discount': '0.0',
         'tax_actual': '2253.2', 'yield': '0.0', 'product_pollute': '0.0', 'discharge_pollute': '0.0',
         'Fe_consume': '-', 'Fe_S_content': '-', 'solid_consume': '-', 'solid_S_content': '-', 'pollutant_yield': '0.0',
         'pollutant_emissions_coeff': '0.0', 'pollutant_yield_equivalent': '0.0',
         'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0', 'emissions_tax': '0.0',
         'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None, 'standard': None,
         'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 15, 'proce_id': 1, 'proce_name': '原料工序', 'production_id': 4, 'production_name': '一期烧结机',
         'outlet_name': '烧结烟气脱硫脱硝废气排放口', 'outlet_type': '7', 'factor': 'so2_2280', 'criterion': None,
         'tax_number': None, 'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口',
         'factor_name': '二氧化硫(2280m³)', 'id': 3289, 'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0,
         'hour_emissions': '-', 'gas_emissions': '12120451.946999999', 'pollutant_concentration': None,
         'pollutant_emissions': '2253.2', 'pollutant_equivalent': '751.0666666666667', 'tax': '2253.2',
         'discount': '0%', 'tax_discount': '0.0', 'tax_actual': '2253.2', 'yield': '0.0', 'product_pollute': '0.0',
         'discharge_pollute': '0.0', 'Fe_consume': '-', 'Fe_S_content': '-', 'solid_consume': '-',
         'solid_S_content': '-', 'pollutant_yield': '0.0', 'pollutant_emissions_coeff': '0.0',
         'pollutant_yield_equivalent': '0.0', 'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0',
         'emissions_tax': '0.0', 'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None,
         'standard': None, 'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 15, 'proce_id': 1, 'proce_name': '原料工序', 'production_id': 5, 'production_name': '二期烧结机',
         'outlet_name': '烧结烟气脱硫脱硝废气排放口', 'outlet_type': '7', 'factor': 'so2_2280', 'criterion': None,
         'tax_number': None, 'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口',
         'factor_name': '二氧化硫(2280m³)', 'id': 3289, 'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0,
         'hour_emissions': '-', 'gas_emissions': '12120451.946999999', 'pollutant_concentration': None,
         'pollutant_emissions': '2253.2', 'pollutant_equivalent': '751.0666666666667', 'tax': '2253.2',
         'discount': '0%', 'tax_discount': '0.0', 'tax_actual': '2253.2', 'yield': '0.0', 'product_pollute': '0.0',
         'discharge_pollute': '0.0', 'Fe_consume': '-', 'Fe_S_content': '-', 'solid_consume': '-',
         'solid_S_content': '-', 'pollutant_yield': '0.0', 'pollutant_emissions_coeff': '0.0',
         'pollutant_yield_equivalent': '0.0', 'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0',
         'emissions_tax': '0.0', 'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None,
         'standard': None, 'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 17, 'proce_id': 7, 'proce_name': '球团工序', 'production_id': 11, 'production_name': '2#竖炉',
         'outlet_name': '2#竖炉焙烧废气排放口', 'outlet_type': '7', 'factor': 'smoke', 'criterion': None, 'tax_number': None,
         'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口', 'factor_name': '颗粒物(烟尘)', 'id': 3292,
         'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0, 'hour_emissions': '-',
         'gas_emissions': '1062634.6800000002', 'pollutant_concentration': None, 'pollutant_emissions': '376.7',
         'pollutant_equivalent': '125.56666666666666', 'tax': '376.7', 'discount': '0%', 'tax_discount': '0.0',
         'tax_actual': '376.7', 'yield': '0.0', 'product_pollute': '0.0', 'discharge_pollute': '0.0', 'Fe_consume': '-',
         'Fe_S_content': '-', 'solid_consume': '-', 'solid_S_content': '-', 'pollutant_yield': '0.0',
         'pollutant_emissions_coeff': '0.0', 'pollutant_yield_equivalent': '0.0',
         'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0', 'emissions_tax': '0.0',
         'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None, 'standard': None,
         'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 17, 'proce_id': 7, 'proce_name': '球团工序', 'production_id': 11, 'production_name': '2#竖炉',
         'outlet_name': '2#竖炉焙烧废气排放口', 'outlet_type': '7', 'factor': 'nox', 'criterion': None, 'tax_number': None,
         'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口', 'factor_name': '氮氧化物', 'id': 3292,
         'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0, 'hour_emissions': '-',
         'gas_emissions': '1062634.6800000002', 'pollutant_concentration': None, 'pollutant_emissions': '376.7',
         'pollutant_equivalent': '125.56666666666666', 'tax': '376.7', 'discount': '0%', 'tax_discount': '0.0',
         'tax_actual': '376.7', 'yield': '0.0', 'product_pollute': '0.0', 'discharge_pollute': '0.0', 'Fe_consume': '-',
         'Fe_S_content': '-', 'solid_consume': '-', 'solid_S_content': '-', 'pollutant_yield': '0.0',
         'pollutant_emissions_coeff': '0.0', 'pollutant_yield_equivalent': '0.0',
         'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0', 'emissions_tax': '0.0',
         'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None, 'standard': None,
         'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None},
        {'outlet_id': 17, 'proce_id': 7, 'proce_name': '球团工序', 'production_id': 11, 'production_name': '2#竖炉',
         'outlet_name': '2#竖炉焙烧废气排放口', 'outlet_type': '7', 'factor': 'so2', 'criterion': None, 'tax_number': None,
         'tax_item': None, 'source': '在线', 'outlet_type_name': '主要排放口', 'factor_name': '二氧化硫', 'id': 3292,
         'month': '2月', 'monitor_time': '02-01', 'run_time': 0.0, 'hour_emissions': '-',
         'gas_emissions': '1062634.6800000002', 'pollutant_concentration': None, 'pollutant_emissions': '376.7',
         'pollutant_equivalent': '125.56666666666666', 'tax': '376.7', 'discount': '0%', 'tax_discount': '0.0',
         'tax_actual': '376.7', 'yield': '0.0', 'product_pollute': '0.0', 'discharge_pollute': '0.0', 'Fe_consume': '-',
         'Fe_S_content': '-', 'solid_consume': '-', 'solid_S_content': '-', 'pollutant_yield': '0.0',
         'pollutant_emissions_coeff': '0.0', 'pollutant_yield_equivalent': '0.0',
         'pollutant_emissions_equivalent_coeff': '0.0', 'yield_tax': '0.0', 'emissions_tax': '0.0',
         'pollutant_unit': '千克', 'unit_tax': None, 'pollutant_equivalent_number': None, 'standard': None,
         'project_name': None, 'equivalent_number': None, 'month_avg': None, 'month_max': None}], 'row1_sum': 15529.36,
                'row2': [{'outlet_id': 53, 'proce_id': 1, 'proce_name': '原料工序', 'production_id': None,
                          'production_name': None, 'outlet_name': '混匀废气排放口', 'outlet_type': '8', 'factor': 'smoke',
                          'criterion': '1', 'tax_number': '2', 'tax_item': '3', 'source': '系数法',
                          'outlet_type_name': '一般排放口', 'factor_name': '颗粒物(烟尘)'},
                         {'outlet_id': 54, 'proce_id': 1, 'proce_name': '原料工序', 'production_id': None,
                          'production_name': None, 'outlet_name': '转运废气排放口', 'outlet_type': '8', 'factor': 'dust',
                          'criterion': None, 'tax_number': None, 'tax_item': None, 'source': '系数法',
                          'outlet_type_name': '一般排放口', 'factor_name': '颗粒物(粉尘)'},
                         {'outlet_id': 93, 'proce_id': 4, 'proce_name': '轧钢工序', 'production_id': None,
                          'production_name': None, 'outlet_name': '2#棒材热处理炉废气排放口（煤烟）', 'outlet_type': '8',
                          'factor': 'so2_1280', 'criterion': '测试', 'tax_number': '测试', 'tax_item': '测试',
                          'source': '系数法', 'outlet_type_name': '一般排放口', 'factor_name': '二氧化硫(1280m³)'},
                         {'outlet_id': 93, 'proce_id': 4, 'proce_name': '轧钢工序', 'production_id': None,
                          'production_name': None, 'outlet_name': '2#棒材热处理炉废气排放口（煤烟）', 'outlet_type': '8',
                          'factor': 'nox_1280', 'criterion': '测试', 'tax_number': '测试', 'tax_item': '测试',
                          'source': '系数法', 'outlet_type_name': '一般排放口', 'factor_name': '氮氧化物(1280m³)'},
                         {'outlet_id': 93, 'proce_id': 4, 'proce_name': '轧钢工序', 'production_id': None,
                          'production_name': None, 'outlet_name': '2#棒材热处理炉废气排放口（煤烟）', 'outlet_type': '8',
                          'factor': 'smoke', 'criterion': '测试', 'tax_number': '测试', 'tax_item': '测试', 'source': '系数法',
                          'outlet_type_name': '一般排放口', 'factor_name': '颗粒物(烟尘)'},
                         {'outlet_id': 94, 'proce_id': 4, 'proce_name': '轧钢工序', 'production_id': None,
                          'production_name': None, 'outlet_name': '1#高线热处理炉废气排放口（空烟）', 'outlet_type': '8',
                          'factor': 'so2_2280', 'criterion': '1', 'tax_number': '2', 'tax_item': '3', 'source': '系数法',
                          'outlet_type_name': '一般排放口', 'factor_name': '二氧化硫(2280m³)'},
                         {'outlet_id': 94, 'proce_id': 4, 'proce_name': '轧钢工序', 'production_id': None,
                          'production_name': None, 'outlet_name': '1#高线热处理炉废气排放口（空烟）', 'outlet_type': '8',
                          'factor': 'nox_2280', 'criterion': '1', 'tax_number': '2', 'tax_item': '3', 'source': '系数法',
                          'outlet_type_name': '一般排放口', 'factor_name': '氮氧化物(2280m³)'},
                         {'outlet_id': 94, 'proce_id': 4, 'proce_name': '轧钢工序', 'production_id': None,
                          'production_name': None, 'outlet_name': '1#高线热处理炉废气排放口（空烟）', 'outlet_type': '8',
                          'factor': 'smoke', 'criterion': '1', 'tax_number': '2', 'tax_item': '3', 'source': '系数法',
                          'outlet_type_name': '一般排放口', 'factor_name': '颗粒物(烟尘)'}], 'row2_sum': 0, 'row3': (),
                'row3_sum': 0, 'row_sum': 15529.36, 'month': '环境保护税按月计算报表（2）月',
                'title': '税款所属期：自2023年02月01日至2023年02月28日'}}
    wb_tmp = load_workbook(Path(__file__).parent / 'template/excel5.xlsx')
    write_excel_for_template(value=value, wb_tmp=wb_tmp, convert_pic=True, dynamic_method=dynamic_method2())
    wb_tmp.save("./val3.xlsx")

if __name__ == '__main__':
    test3()
    # test2()