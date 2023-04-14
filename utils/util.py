# -*- coding: utf-8 -*-
# !/usr/bin/env python
# @Time    : 2023/2/13 16:28 
# @Author  : mtl
# @Desc    : excel 增加自定义方法执行
# @File    : util.py
# @Software: PyCharm
import collections
import re

from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from simple_export.utils.tool import char_to_num, num_to_pos_char
from openpyxl.cell import Cell

class dynamic_method():
    def __call__(self, *args, **kwargs):
        funcs = args[0]
        for func in funcs:
            if func:
                arg: tuple = args[1:]
                match_value: list = re.findall('\((.+)\)', func)
                if match_value:
                    arg += tuple(str(match_value[0]).split(","))
                    func = str(func).split("(")[0]
                call_function = getattr(self, func)
                call_function(*arg, **kwargs)

    """
    合并单元格向下合并
    """
    def merge_bottom(self, coordinate: str, sheet: Worksheet, pos_mapping: collections.defaultdict, pos: collections.defaultdict ) -> None:
        """
        将指定单元格下方的所有单元格合并为一个单元格。
        :param coordinate: 单元格的坐标，例如 "A1"。
        :param sheet: 工作表对象。
        :param pos_mapping: 单元格坐标到行列号的映射。
        :return: 无返回值。
        """
        col, row = char_to_num(coordinate)
        col -= 1
        row -= 1
        p = pos_mapping[row]
        start_row, end_row = p[0], p[-1]
        start = start_row - 1
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for i in range(start_row, end_row):
            current_cell, prev_cell = pos[i][col], pos[i - 1][col]
            if isinstance(current_cell, Cell) and isinstance(prev_cell, Cell) and current_cell.value != prev_cell.value:
                if i > start + 1:
                    sheet.merge_cells(start_row=start + 1, start_column=col + 1, end_column=col + 1, end_row=i)
                    sheet[num_to_pos_char((col + 1, start + 1))].alignment = align
                start = i
        if end_row - start > 1:
            sheet.merge_cells(start_row=start + 1, start_column=col + 1, end_column=col + 1, end_row=end_row)
            sheet[num_to_pos_char((col + 1, start + 1))].alignment = align

    """
    合并单元格向左合并
    """
    def merge_left(self, coordinate: str, sheet: Worksheet, pos_mapping: collections.defaultdict, pos: collections.defaultdict ) -> None:
        """
        将指定单元格下方的所有单元格合并为一个单元格。
        :param coordinate: 单元格的坐标，例如 "A1"。
        :param sheet: 工作表对象。
        :param pos_mapping: 单元格坐标到行列号的映射。
        :return: 无返回值。
        """
        col, row = char_to_num(coordinate)
        col -= 1
        row -= 1
        p = pos_mapping[row]
        c = pos_mapping[col]
        start_row, end_row = p[0], p[-1]
        start = 0
        end_col = col + 1
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for i in range(start_row, end_row):
            for j in range(1, end_col):
                current_cell, prev_cell = pos[i][col], pos[i][col - j]
                if isinstance(current_cell, Cell) and isinstance(prev_cell, Cell) and current_cell.value != prev_cell.value:
                    if end_col - j + 1 != i+1:
                        sheet.merge_cells(start_row=i+1, start_column=end_col - j + 1, end_column=end_col, end_row=i+1)
                        sheet[num_to_pos_char((end_col - j + 1, i+1))].alignment = align
                        break
                start = j
            if end_col - start <= 1:
                sheet.merge_cells(start_row=i+1, start_column=1, end_column=end_col, end_row=i+1)
                sheet[num_to_pos_char((1, i+1))].alignment = align